from io import BytesIO

from openpyxl import load_workbook


def auth_headers(access_token: str, client_app: str = "employee") -> dict[str, str]:
    return {
        "Authorization": f"Bearer {access_token}",
        "X-Client-App": client_app,
    }


def register(client, client_app: str, fingerprint: str) -> dict:
    response = client.post(
        "/api/v1/devices/register",
        headers={"X-Client-App": client_app},
        json={"fingerprint_hash": fingerprint * 64},
    )
    assert response.status_code == 200
    return response.json()


def assign_role(client, actor: dict, target: dict, role: str) -> None:
    response = client.put(
        f"/api/v1/employee/devices/{target['device_id']}/role",
        headers=auth_headers(actor["access_token"]),
        json={"role": role},
    )
    assert response.status_code == 200


def test_chief_downloads_excel_report_with_required_sheets(client):
    observer = register(client, "eyewitness", "a")
    chief = register(client, "employee", "b")

    message = client.post(
        "/api/v1/messages",
        headers=auth_headers(observer["access_token"], "eyewitness"),
        json={"message_type": "TEXT", "text": "report me"},
    )
    ban = client.post(
        f"/api/v1/employee/devices/{observer['device_id']}/ban",
        headers=auth_headers(chief["access_token"]),
    )
    answer = client.post(
        "/api/v1/messages",
        headers=auth_headers(chief["access_token"]),
        json={
            "observer_device_id": observer["device_id"],
            "message_type": "TEXT",
            "text": "answer",
        },
    )
    report = client.get(
        "/api/v1/employee/reports/excel",
        headers=auth_headers(chief["access_token"]),
    )

    assert message.status_code == 200
    assert ban.status_code == 200
    assert answer.status_code == 200
    assert report.status_code == 200
    assert report.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert report.headers["content-disposition"] == (
        'attachment; filename="gibdd-report.xlsx"'
    )

    workbook = load_workbook(BytesIO(report.content))
    assert workbook.sheetnames == ["Bans", "Roles", "Messages"]
    assert [cell.value for cell in workbook["Bans"][1]] == [
        "ban_id",
        "observer_device_id",
        "issued_by_device_id",
        "started_at",
        "ends_at",
        "ban_number",
        "is_active",
    ]
    assert [cell.value for cell in workbook["Roles"][1]] == [
        "event_id",
        "actor_device_id",
        "target_device_id",
        "action",
        "role",
        "created_at",
    ]
    assert [cell.value for cell in workbook["Messages"][1]] == [
        "sender_device_id",
        "message_count",
        "text_count",
        "media_count",
        "static_location_count",
        "live_location_count",
        "delivered_count",
        "first_created_at",
        "last_created_at",
    ]
    assert workbook["Bans"].max_row == 2
    assert workbook["Messages"].max_row == 3

    rows = {
        row[0]: row[1:]
        for row in workbook["Messages"].iter_rows(min_row=2, values_only=True)
    }
    assert rows[observer["device_id"]][:6] == (1, 1, 0, 0, 0, 0)
    assert rows[chief["device_id"]][:6] == (1, 1, 0, 0, 0, 0)


def test_admin_cannot_download_excel_report(client):
    chief = register(client, "employee", "c")
    admin = register(client, "employee", "d")
    assign_role(client, chief, admin, "ADMIN")

    report = client.get(
        "/api/v1/employee/reports/excel",
        headers=auth_headers(admin["access_token"]),
    )

    assert report.status_code == 403
    assert report.json()["detail"] == "Reports are allowed only for CHIEF"
