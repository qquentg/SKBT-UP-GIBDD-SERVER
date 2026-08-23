from io import BytesIO
from copy import copy

from fastapi import HTTPException, status
from openpyxl import Workbook

from app.models.ban import Ban
from app.models.device import Device
from app.models.message import Message
from app.models.role_event import RoleEvent
from app.schemas.device import DeviceRole
from app.schemas.messages import MessageType
from app.services.bans import ban_number, is_ban_active


def require_report_access(actor: Device) -> None:
    if actor.current_role != DeviceRole.CHIEF.value:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Reports are allowed only for CHIEF",
        )


def generate_excel_report() -> bytes:
    workbook = Workbook()
    _fill_bans_sheet(workbook.active)
    _fill_roles_sheet(workbook.create_sheet("Roles"))
    _fill_messages_sheet(workbook.create_sheet("Messages"))

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _fill_bans_sheet(sheet) -> None:
    sheet.title = "Bans"
    _append_header(
        sheet,
        [
            "ban_id",
            "observer_device_id",
            "issued_by_device_id",
            "started_at",
            "ends_at",
            "ban_number",
            "is_active",
        ],
    )
    for ban in Ban.select().order_by(Ban.started_at.asc(), Ban.id.asc()):
        sheet.append(
            [
                str(ban.id),
                str(ban.observer_device_id),
                str(ban.issued_by_device_id),
                _datetime_value(ban.started_at),
                _datetime_value(ban.ends_at),
                ban_number(ban),
                is_ban_active(ban),
            ]
        )


def _fill_roles_sheet(sheet) -> None:
    _append_header(
        sheet,
        [
            "event_id",
            "actor_device_id",
            "target_device_id",
            "action",
            "role",
            "created_at",
        ],
    )
    for event in RoleEvent.select().order_by(RoleEvent.created_at.asc(), RoleEvent.id.asc()):
        sheet.append(
            [
                str(event.id),
                str(event.actor_device_id) if event.actor_device_id is not None else None,
                str(event.target_device_id),
                event.action,
                event.role,
                _datetime_value(event.created_at),
            ]
        )


def _fill_messages_sheet(sheet) -> None:
    _append_header(
        sheet,
        [
            "sender_device_id",
            "message_count",
            "text_count",
            "media_count",
            "static_location_count",
            "live_location_count",
            "delivered_count",
            "first_created_at",
            "last_created_at",
        ],
    )
    summaries = _message_summaries_by_sender()
    for sender_device_id in sorted(summaries):
        summary = summaries[sender_device_id]
        sheet.append(
            [
                sender_device_id,
                summary["message_count"],
                summary["text_count"],
                summary["media_count"],
                summary["static_location_count"],
                summary["live_location_count"],
                summary["delivered_count"],
                _datetime_value(summary["first_created_at"]),
                _datetime_value(summary["last_created_at"]),
            ]
        )


def _message_summaries_by_sender() -> dict[str, dict]:
    summaries: dict[str, dict] = {}
    for message in Message.select().order_by(Message.created_at.asc(), Message.id.asc()):
        sender_device_id = str(message.sender_device_id)
        summary = summaries.setdefault(
            sender_device_id,
            {
                "message_count": 0,
                "text_count": 0,
                "media_count": 0,
                "static_location_count": 0,
                "live_location_count": 0,
                "delivered_count": 0,
                "first_created_at": message.created_at,
                "last_created_at": message.created_at,
            },
        )
        summary["message_count"] += 1
        summary["delivered_count"] += int(message.delivered_at is not None)
        summary["first_created_at"] = min(
            summary["first_created_at"], message.created_at
        )
        summary["last_created_at"] = max(summary["last_created_at"], message.created_at)

        if message.message_type == MessageType.TEXT.value:
            summary["text_count"] += 1
        elif message.message_type == MessageType.MEDIA.value:
            summary["media_count"] += 1
        elif message.message_type == MessageType.STATIC_LOCATION.value:
            summary["static_location_count"] += 1
        elif message.message_type == MessageType.LIVE_LOCATION.value:
            summary["live_location_count"] += 1

    return summaries


def _append_header(sheet, values: list[str]) -> None:
    sheet.append(values)
    for cell in sheet[1]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font


def _datetime_value(value):
    return value.isoformat() if value is not None else None
